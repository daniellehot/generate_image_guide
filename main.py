#from copy import copy
#from tkinter import N
from csv import excel_tab
import email
from email.mime import image
from re import M
import numpy as np
import random
import os
#import csv
import openpyxl

min_number_of_specimens = 5
max_number_of_specimens = 15
stat = [0, 0, 0, 0] #cod, pollock, haddock, whitting
number_of_images = 250
excel_file = ("guide.xlsx")
empty_row = ["------------------" for i in range(26)]


def generate_species_and_specimens():
    #species = ["cod", "pollock", "haddock", "whitting", "other"]
    options = ["cod", "pollock", "haddock", "whitting"]
    sum = random.randint(min_number_of_specimens, max_number_of_specimens)
    #print(sum)

    arr = np.random.randint(100, size=len(options))
    arr = np.divide(arr, np.sum(arr))
    #print(arr)
    
    specimens = np.around(arr*sum)
    #print(specimens)

    random.shuffle(options)
    #print(options)

    species_order = []
    for (fish, number) in zip(options, specimens):
        for i in range(int(number)):
            species_order.append(fish)
    #print(species_order)

    return specimens, options, species_order
    
    
def generate_mapping(species_order):
    options = ['LT', 'LC', 'LB', 'CT', 'CC', 'CB', 'RT', 'RC', 'RB']
    len_options = len(options)
    diff = len(species_order) - len(options)
    if diff > 0:
        for i in range(diff):
            options.append(options[random.randint(0, len_options-1)])
    if diff < 0:
        for i in range(abs(diff)):
             options.pop(random.randint(0, len(options)-1))
    
    random.shuffle(options)
    #print(options)
    return options


def get_statistics(specimens, species_order):
    species = ["cod", "pollock", "haddock", "whitting"]
    global stat
    for (number, fish) in zip(specimens, species_order):
        idx = species.index(fish)
        stat[idx] += number 


def  generate_rotations(locations):
    return [random.randint(1,12) for i in range(len(locations))]


def save_data(path, row):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.append(row)
    wb.save(filename="guide.xlsx")


def reset_guide(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.delete_rows(1, ws.max_row+1)
    wb.save(filename="guide.xlsx")


if __name__=="__main__":
    #Check if the file is empty. If not, clean it
    if os.stat(excel_file).st_size != 0:
        reset_guide(excel_file)

    for i in range(number_of_images):
        specimens, species_order, species_order_long = generate_species_and_specimens()
        mapping = generate_mapping(species_order_long)
        rotations = generate_rotations(mapping)
        statistics = get_statistics(specimens, species_order)
        
        print(specimens)
        print(species_order)
        print(mapping)
        print(rotations)
        print("%%%%%%%%%%%%%%%%%%%")
        for (fish, number) in zip(species_order, specimens):
            number = int(number)
            row = []
            row.append(i+1)
            row.append(fish)
            row.append(number)
            row.append("||")
            for j in range(number):
                row.append(mapping[j])
                row.append(rotations[j])
                row.append("||")
            del mapping[:number]
            del rotations[:number]
            print(row)
            save_data(excel_file, row)
        save_data(excel_file, empty_row)
            
            
        #save_data(excel_file, i, specimens, species_order, mapping, rotations)
    #print(stat)

    for value in stat:
        temp = value/sum(stat)*100
        #print(temp)